#!/usr/bin/env python3
"""
VIMS_REPLAY_SIMULATOR

Simulateur des 20 capteurs CAT 994F1 P1+P2 EXACTEMENT identiques aux
exports VIMS reels de l'utilisateur (fichier
"Paramètres+Diagnostique_*.xlsx" produit par CAT Diagnostic Reporter).

Sortie au choix :
  --csv         : fichier CSV au pas de 1 s (timestamps + 20 colonnes)
  --xlsx        : fichier Excel au format VIMS (snapshot 2 min, min/avg/max)
  --post URL    : envoi temps reel POST a un backend (defaut MineAssist)

Le simulateur reproduit :
  - Cycle moteur 6 phases (idle / reprise / charge / pleine charge / decharge / retour)
  - Couplage realiste entre Regime moteur, Pression huile, Temperatures
  - Boucle thermique sur Temperature liquide refroidissement (modele ventilo)
  - Modele hydraulique (Pression pompe principale)
  - Cycles de transmission (Lock-up + Impeller)
  - Tension alternateur stable autour de 27 V
  - Compresseur d'air avec hysteresis
  - Bruit de mesure realiste sur chaque canal

Usage :
  python vims_replay_simulator.py --duration 600 --csv sortie.csv
  python vims_replay_simulator.py --duration 1800 --xlsx export_vims.xlsx
  python vims_replay_simulator.py --duration 600 --fault ventilo_hs --t-fault 60
  python vims_replay_simulator.py --duration 300 --post http://127.0.0.1:8000/sim/ingest

PFE MineAssist - OCP Benguerir - mai 2026
"""
from __future__ import annotations
import argparse
import json
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import numpy as np

ROOT = Path(__file__).resolve().parent
SENSORS_FILE = ROOT / "vims_sensors.json"


# =============================================================================
# 1. CHARGEMENT DES CAPTEURS
# =============================================================================
def load_sensors() -> dict:
    with open(SENSORS_FILE, encoding="utf-8") as f:
        return json.load(f)


# =============================================================================
# 2. MODELE MOTEUR : CYCLE CHARGEUR EN 6 PHASES (60 s par cycle complet)
# =============================================================================
def cycle_chargeur(t: float) -> dict:
    """Renvoie un dict des etats moteur a l'instant t (s).

    Cycle de 60 s pour un chargeur 994F :
      0-10 s  : idle (regime bas, pas de charge hydraulique)
      10-15 s : reprise moteur (montee en regime)
      15-30 s : creusage (charge max sur les verins)
      30-40 s : pleine charge (regime nominal max)
      40-50 s : retour vide (decharge hydraulique)
      50-60 s : ralenti retour (idle de nouveau)
    """
    phase = (t % 60.0) / 60.0
    if phase < 10/60:
        # idle
        rpm_norm = 0.10  # ~750 rpm
        load_norm = 0.05
        hyd_norm = 0.02
    elif phase < 15/60:
        # reprise
        x = (phase - 10/60) / (5/60)
        rpm_norm = 0.10 + x * 0.55  # 750 -> 1450 rpm
        load_norm = 0.05 + x * 0.35
        hyd_norm = 0.02 + x * 0.55
    elif phase < 30/60:
        # creusage : pic haute pression hydraulique (~17 MPa)
        # Seuil OCP 6 : pendant rpm > 1500, P_hyd doit etre >= 15000 kPa
        rpm_norm = 0.85
        load_norm = 0.90
        hyd_norm = 0.65    # P = 30 + 26000*0.65 = 16930 kPa ✓ seuil OCP
    elif phase < 40/60:
        # pleine charge : maintien hydraulique a regime nominal max
        rpm_norm = 0.95
        load_norm = 0.95
        hyd_norm = 0.62    # P = 30 + 26000*0.62 = 16150 kPa ✓ seuil OCP
    elif phase < 50/60:
        # retour vide (pression baisse)
        rpm_norm = 0.55
        load_norm = 0.40
        hyd_norm = 0.10
    else:
        # idle retour
        rpm_norm = 0.15
        load_norm = 0.15
        hyd_norm = 0.04
    return {"rpm_norm": rpm_norm, "load_norm": load_norm, "hyd_norm": hyd_norm}


# =============================================================================
# 3. ETAT THERMIQUE : modele 2 noeuds (eau + metal radiateur) + ventilo
# =============================================================================
@dataclass
class ThermalState:
    T_eau: float = 75.0       # degC
    T_met: float = 70.0       # degC
    T_huile_dir: float = 22.0
    T_huile_frein: float = 25.0
    T_PTO: float = 22.0
    T_essieu_av: float = 22.0
    T_essieu_ar: float = 22.0
    T_conv: float = 30.0
    T_echap_d: float = 30.0
    T_echap_g: float = 30.0
    fan_state: bool = False

    def step(self, dt: float, load_norm: float, T_amb: float = 30.0,
             ventilo_eff: float = 1.0, k_em: float = 1200.0,
             C_eau: float = 50e3, C_met: float = 25e3,
             k_off: float = 200.0, k_on: float = 1800.0):
        """Avance les noeuds thermiques d'un pas dt (s).

        ventilo_eff : 0..1, multiplie l'efficacite k_on (en cas de defaut).
        """
        # Hysteresis ventilo 82/85 (seuils MineAssist)
        if not self.fan_state and self.T_eau >= 85.0:
            self.fan_state = True
        elif self.fan_state and self.T_eau <= 82.0:
            self.fan_state = False

        if self.fan_state:
            k_air = k_off + (k_on - k_off) * ventilo_eff
        else:
            k_air = k_off

        # Puissance dissipee dans l'eau, fonction de la charge moteur
        P_diss = 8000.0 + 22000.0 * load_norm  # 8 kW idle -> 30 kW pleine charge

        q_em = k_em * (self.T_eau - self.T_met)
        q_air = k_air * (self.T_met - T_amb)
        self.T_eau += dt * (P_diss - q_em) / C_eau
        self.T_met += dt * (q_em - q_air) / C_met

        # Autres temperatures (1er ordre vers cible_charge)
        # PTO avant
        T_pto_target = 22.0 + 65.0 * load_norm
        self.T_PTO += dt * (T_pto_target - self.T_PTO) / 60.0
        # Huile direction
        T_dir_target = 22.0 + 50.0 * load_norm
        self.T_huile_dir += dt * (T_dir_target - self.T_huile_dir) / 90.0
        # Huile freinage (legerement plus chaud)
        T_frein_target = 25.0 + 60.0 * load_norm
        self.T_huile_frein += dt * (T_frein_target - self.T_huile_frein) / 75.0
        # Essieux (lent)
        T_essav_target = 22.0 + 40.0 * load_norm
        self.T_essieu_av += dt * (T_essav_target - self.T_essieu_av) / 120.0
        T_essar_target = 22.0 + 45.0 * load_norm
        self.T_essieu_ar += dt * (T_essar_target - self.T_essieu_ar) / 120.0
        # Sortie convertisseur (vmoy reel ~ 93 degC)
        T_conv_target = 50.0 + 65.0 * load_norm
        self.T_conv += dt * (T_conv_target - self.T_conv) / 45.0
        # Echappement (rapide, suit le regime ~ 4eme puissance)
        T_echap_target = 30.0 + 540.0 * (load_norm ** 0.7)
        self.T_echap_d += dt * (T_echap_target - self.T_echap_d) / 12.0
        self.T_echap_g += dt * (T_echap_target - 8 - self.T_echap_g) / 12.0


# =============================================================================
# 4. SIMULATEUR PRINCIPAL
# =============================================================================
@dataclass
class SimulatorConfig:
    duration_s: float = 600.0
    dt: float = 1.0
    fault: str = ""           # ventilo_hs / surchauffe_progressive / fuite_huile / niveau_bas
    t_fault: float = 60.0     # instant d'apparition du defaut (s)
    speed: float = 1.0        # 1.0 = temps reel, >1 plus rapide
    seed: int = 42
    engin: str = "994 F1"


def simulate(cfg: SimulatorConfig) -> tuple[list, list]:
    """Lance la simulation et renvoie (timestamps, samples).

    samples est une liste de dicts avec une cle par capteur (nom exact VIMS).
    """
    rng = np.random.default_rng(cfg.seed)
    sensors = load_sensors()

    # Etat
    therm = ThermalState()
    P_huile_moteur = 450.0
    P_air = 750.0   # kPa (au milieu de la plage OCP 600..900)
    air_charging = False
    P_imp = 1200.0
    I_imp = 60.0
    I_lock = 60.0
    V_sys = 27_000.0  # mV
    debit_eau = 1     # 0/1
    niveau_huile_low = 127  # constant a 127, chute si fuite

    # Timestamps simules (debut maintenant)
    t0 = datetime.now().replace(microsecond=0)
    timestamps: list[datetime] = []
    samples: list[dict] = []

    n = int(cfg.duration_s / cfg.dt) + 1
    for i in range(n):
        t = i * cfg.dt
        ts = t0 + timedelta(seconds=int(t))
        # ===== Cycle moteur =====
        c = cycle_chargeur(t)
        rpm_norm = c["rpm_norm"]
        load_norm = c["load_norm"]
        hyd_norm = c["hyd_norm"]

        # ===== Defauts =====
        ventilo_eff = 1.0
        load_extra = 0.0
        if cfg.fault and t >= cfg.t_fault:
            if cfg.fault == "ventilo_hs":
                # Degradation lineaire sur 600 s
                ventilo_eff = max(0.0, 1.0 - (t - cfg.t_fault) / 600.0)
            elif cfg.fault == "surchauffe_progressive":
                load_extra = min(0.4, (t - cfg.t_fault) / 300.0 * 0.4)
            elif cfg.fault == "niveau_bas":
                # tombe a 0 progressivement
                if t - cfg.t_fault > 30:
                    niveau_huile_low = max(1, int(127 - (t - cfg.t_fault - 30) * 2))
            elif cfg.fault == "fuite_huile":
                P_huile_moteur_target = max(50, 450 - (t - cfg.t_fault) / 5.0 * 1.0)
                # P_huile_moteur descend tres lentement
                P_huile_moteur += (P_huile_moteur_target - P_huile_moteur) * 0.05

        # ===== Thermique (1 pas) =====
        therm.step(dt=cfg.dt, load_norm=load_norm + load_extra,
                   ventilo_eff=ventilo_eff)

        # ===== Regime moteur =====
        rpm = 750.0 + 1010.0 * rpm_norm + rng.normal(0, 8)
        rpm = float(np.clip(rpm, 670, 1780))

        # ===== Pression huile moteur (correlee au regime + bruit) =====
        if cfg.fault != "fuite_huile":
            P_huile_target = 250.0 + 360.0 * rpm_norm
            P_huile_moteur += (P_huile_target - P_huile_moteur) * 0.20 + rng.normal(0, 4)
        P_huile_moteur = float(np.clip(P_huile_moteur, 30, 620))

        # ===== Pression hydraulique principale (cycle creusage) =====
        # Reference reel : vmoy=7378 kPa (74 bar), vmax=29742 kPa
        # Modele : pression de fond 30 kPa (idle) + composante load
        P_hyd = 30.0 + 26000.0 * hyd_norm + rng.normal(0, 250)
        P_hyd = float(np.clip(P_hyd, 0, 30000))
        # Saturation soupape de tarage 28 MPa = 28000 kPa
        if P_hyd > 28000:
            P_hyd = 28000 + rng.normal(0, 80)

        # ===== Compresseur d'air (hysteresis dans plage OCP 600..900 kPa) =====
        # Seuils OCP : alerte si P_air sort de [600, 900]. On regle l'hysteresis
        # interne pour rester confortablement dans la plage.
        if not air_charging and P_air <= 700:
            air_charging = True
        elif air_charging and P_air >= 870:
            air_charging = False
        if air_charging:
            P_air += 4.0 * cfg.dt
        else:
            P_air -= 0.6 * cfg.dt
        P_air = float(np.clip(P_air, 620, 890))

        # ===== Pression / courant embrayage impeller =====
        # Seuil OCP 14 : a rpm >= 1510, P_imp regule dans [1860, 1870] kPa
        # En creusage / pleine charge (rpm_norm >= 0.85), regulation serree CAT
        if rpm_norm >= 0.84:  # phases creusage + pleine charge
            # Regulation directe a 1865 kPa, bruit serre + clip dans [1861, 1869]
            P_imp = 1865.0 + rng.normal(0, 1.0)
            P_imp = float(np.clip(P_imp, 1861.0, 1869.0))
        else:
            # Regime bas : montee transitoire avec rpm_norm
            P_imp_target = 100.0 + 2200.0 * rpm_norm
            P_imp += (P_imp_target - P_imp) * 0.1 + rng.normal(0, 30)
        P_imp = float(np.clip(P_imp, 30, 2280))
        # Courant impeller (% mais vmax 241 dans donnees -> echelle CAT)
        I_imp_target = 35 + 200 * rpm_norm
        I_imp += (I_imp_target - I_imp) * 0.3 + rng.normal(0, 2)
        I_imp = float(np.clip(I_imp, 30, 245))
        # Courant lock-up (couplage aux changements de phase)
        I_lock_target = 55 + 180 * load_norm
        I_lock += (I_lock_target - I_lock) * 0.25 + rng.normal(0, 2)
        I_lock = float(np.clip(I_lock, 50, 245))

        # ===== Regime sortie convertisseur =====
        rpm_out = 30 + 2100 * rpm_norm + rng.normal(0, 12)
        rpm_out = float(np.clip(rpm_out, 20, 2150))

        # ===== Tension systeme (24 V : 27000 mV nominal) =====
        # Petites chutes en pleine charge
        V_target = 27200.0 - 200 * load_norm
        V_sys += (V_target - V_sys) * 0.10 + rng.normal(0, 25)
        V_sys = float(np.clip(V_sys, 24500, 27900))

        # ===== Debit liquide =====
        if cfg.fault == "ventilo_hs" and ventilo_eff < 0.05:
            debit_eau = 0    # impossible normalement, ici defaut grave
        else:
            debit_eau = 1

        # ===== Sample = dict {nom_VIMS : valeur} =====
        sample = {
            "CH994.P1.Régime moteur":                            round(rpm, 1),
            "CH994.P1.Pression huile moteur":                    round(P_huile_moteur, 1),
            "CH994.P1.Température liquide refroidissement":      round(therm.T_eau, 2),
            "CH994.P1.Température sortie convertisseur":         round(therm.T_conv, 2),
            "CH994.P1.Température échappement Droit":            round(therm.T_echap_d, 1),
            "CH994.P1.Température échappement gauche":           round(therm.T_echap_g, 1),
            "CH994.P1.Température huile direction":              round(therm.T_huile_dir, 2),
            "CH994.P1.Température huile freinage":               round(therm.T_huile_frein, 2),
            "CH994.P1.Température PTO avant":                    round(therm.T_PTO, 2),
            "CH994.P1.Débit liquide refroidissement":            int(debit_eau),
            "CH994.P1.Niveau huile moteur bas":                  int(niveau_huile_low),
            "CH994.P2.Pression pompe hydraulique principale":    round(P_hyd, 0),
            "CH994.P2.Pression d’air au réservoir":              round(P_air, 0),
            "CH994.P2.Pression embrayage impeller":              round(P_imp, 0),
            "CH994.P2.Courant embrayage impeller":               round(I_imp, 1),
            "CH994.P2.Courant embrayage Lock-up":                round(I_lock, 1),
            "CH994.P2.Régime sortie convertisseur":              round(rpm_out, 0),
            "CH994.P2.Température Essieux avant":                round(therm.T_essieu_av, 2),
            "CH994.P2.Température essieux arrière":              round(therm.T_essieu_ar, 2),
            "CH994.P2.Tension électrique de système":            int(V_sys),
        }
        timestamps.append(ts)
        samples.append(sample)

        # Heartbeat console
        if i % 60 == 0:
            print(f"  t={t:5.0f}s  RPM={rpm:.0f}  T_eau={therm.T_eau:5.1f}degC  "
                  f"P_hyd={P_hyd/100:.1f}bar  fan={'ON ' if therm.fan_state else 'OFF'}",
                  flush=True)

    return timestamps, samples


# =============================================================================
# 5. EXPORT CSV (1 Hz)
# =============================================================================
def export_csv(timestamps: list, samples: list, path: Path,
               sensors: dict, engin: str = "994 F1"):
    """CSV format brut : 1 ligne par snapshot, 1 colonne par capteur."""
    import csv
    columns = ["Heure"] + [s["nom"] for s in sensors["capteurs"]]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(columns)
        for ts, sample in zip(timestamps, samples):
            row = [ts.strftime("%Y-%m-%d %H:%M:%S")]
            for s in sensors["capteurs"]:
                row.append(sample.get(s["nom"], ""))
            writer.writerow(row)
    print(f"[ok] CSV ecrit -> {path} ({len(samples)} lignes x {len(columns)} cols)")


# =============================================================================
# 6. EXPORT EXCEL (format VIMS exact : aggregation 2 min min/avg/max)
# =============================================================================
def export_xlsx_vims_format(timestamps: list, samples: list, path: Path,
                            sensors: dict, engin: str = "994 F1"):
    """Reproduit EXACTEMENT le format VIMS du fichier d'origine :
       - lignes d'en-tete (Enterprise, Engin, Intervalle, Paramètres Diagnostic)
       - une ligne par capteur x snapshot 2 min, avec min/avg/max
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport de diagnostic paramètre"

    # En-tete
    bold = Font(bold=True, size=14)
    ws["A1"] = "Rapport de diagnostic paramètres"
    ws["A1"].font = bold
    ws["A3"] = "Enterprise"
    ws["B3"] = "Benguérir"
    ws["A4"] = "Engin"
    ws["B4"] = engin
    ts_min = timestamps[0].strftime("%d.%m.%Y %H:%M:%S")
    ts_max = timestamps[-1].strftime("%d.%m.%Y %H:%M:%S")
    ws["A5"] = "Intervalle"
    ws["B5"] = f"{ts_min} - {ts_max}"
    ws["A6"] = "Paramètres Diagnostic"
    ws["B6"] = f"{len(sensors['capteurs'])} objet"

    # Headers ligne 9
    headers = ["Engin", "Paramètres Diagnostic", "Code", "Heure",
               "Valeur minimale", "Valeur moyenne", "Valeur maximale",
               "Unité de mesure", "Fonctionnement du capteur"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=9, column=j, value=h)
        c.font = Font(bold=True)

    # Aggregation par fenetre 2 min sur chaque capteur
    window_s = 120
    n = len(timestamps)
    nwin = max(1, n // window_s)

    row = 10
    for j in range(nwin):
        i0 = j * window_s
        i1 = min(n, (j + 1) * window_s)
        if i1 - i0 < 1:
            continue
        ts_snap = timestamps[i1 - 1]   # comme VIMS : timestamp = fin de fenetre
        for s in sensors["capteurs"]:
            nom = s["nom"]
            unite = s["unite"]
            code = s["code"]
            vals = [sample[nom] for sample in samples[i0:i1]]
            # Pour les binaires, valeur entiere
            if s["type"] in ("binaire", "compteur_defaut"):
                vmin = int(min(vals))
                vmoy = int(round(sum(vals) / len(vals)))
                vmax = int(max(vals))
            else:
                vmin = int(round(min(vals)))
                vmoy = int(round(sum(vals) / len(vals)))
                vmax = int(round(max(vals)))
            ws.cell(row=row, column=1, value=engin)
            ws.cell(row=row, column=2, value=nom)
            ws.cell(row=row, column=3, value=code)
            ws.cell(row=row, column=4, value=ts_snap)
            ws.cell(row=row, column=5, value=vmin)
            ws.cell(row=row, column=6, value=vmoy)
            ws.cell(row=row, column=7, value=vmax)
            ws.cell(row=row, column=8, value=unite if unite else None)
            ws.cell(row=row, column=9, value="Oui")
            row += 1

    # Largeur colonnes
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["D"].width = 20

    wb.save(path)
    print(
        f"[ok] Excel VIMS ecrit -> {path} ({nwin} snapshots x {len(sensors['capteurs'])} capteurs)")


# =============================================================================
# 7. POST TEMPS REEL (compatible /sim/ingest schema IngestRequest)
# =============================================================================
def _cycle_phase_replay(t: float) -> str:
    """Mappe la phase courante du cycle 60 s vers les libelles attendus
    par le backend (mineassist_live_simulator.py utilise les memes labels)."""
    phase = (t % 60.0) / 60.0
    if phase < 10/60:
        return "approche"      # idle 0-10s
    if phase < 15/60:
        return "levage"        # reprise 10-15s
    if phase < 30/60:
        return "creusage"      # 15-30s
    if phase < 40/60:
        return "pleine charge"  # 30-40s
    if phase < 50/60:
        return "vidage"        # 40-50s
    return "retour"            # 50-60s


def post_realtime(timestamps: list, samples: list, url: str,
                  speed: float = 1.0, engin: str = "994F1",
                  fault: str = "") -> dict:
    """Envoie chaque sample en HTTP POST au backend MineAssist.

    Le payload respecte le schema `IngestRequest` du sim_router (FastAPI) :
      {
        "engin": "994F1",
        "horodatage": "...",
        "mesures": [{"parametre": "CH994...", "valeur": ..., "unite": ...}, ...],
        "cycle_phase": "creusage|pleine charge|...",
        "defaut_actif": "ventilo_hs|fuite_huile|..." | None
      }

    Renvoie un dict de stats (n_ok, n_err, latency_ms_avg, last_status, ...).
    """
    import urllib.request

    # Map nom_capteur_VIMS -> unite (pour aider le backend / dashboard)
    UNITE_PAR_CAPTEUR = {
        "Régime moteur": "Tr/min",
        "Régime sortie convertisseur": "Tr/min",
        "Pression huile moteur": "kPa",
        "Pression pompe hydraulique principale": "kPa",
        "Pression d\u2019air au réservoir": "kPa",
        "Pression embrayage impeller": "kPa",
        "Courant embrayage impeller": "A",
        "Courant embrayage Lock-up": "A",
        "Tension électrique de système": "mV",
        "Niveau huile moteur bas": "bool",
        "Débit liquide refroidissement": "bool",
    }
    DEFAULT_UNIT_TEMP = "\u00b0C"  # tout ce qui commence par "Température"

    def _unite_for(nom_brut: str) -> str:
        # nom_brut = "CH994.P1.Pression pompe hydraulique principale"
        clean = nom_brut.split(".", 2)[-1] if "." in nom_brut else nom_brut
        if clean in UNITE_PAR_CAPTEUR:
            return UNITE_PAR_CAPTEUR[clean]
        if clean.lower().startswith("temp"):
            return DEFAULT_UNIT_TEMP
        return ""

    print(f"[post] envoi temps reel -> {url}  (speed={speed}x, fault={fault or 'aucun'})")
    n = len(samples)
    t_start = time.time()
    n_ok = 0
    n_err = 0
    last_status = ""
    latencies_ms: list[float] = []

    t0_sim = timestamps[0] if timestamps else datetime.now()

    for i, (ts, sample) in enumerate(zip(timestamps, samples)):
        # Phase cycle (modulo 60 s)
        try:
            t_in_run = (ts - t0_sim).total_seconds()
        except Exception:
            t_in_run = float(i)
        phase = _cycle_phase_replay(t_in_run)

        # defaut_actif : declenche apres t-fault (gere par le simulator,
        # ici on transmet uniquement le nom)
        defaut_actif: Optional[str] = fault if fault else None

        # Construit la liste mesures dans le schema IngestRequest
        mesures = []
        for nom_capteur, val in sample.items():
            try:
                v = float(val)
            except (TypeError, ValueError):
                continue
            mesures.append({
                "parametre": nom_capteur,
                "valeur":    v,
                "unite":     _unite_for(nom_capteur),
            })

        body = {
            "engin":         engin,
            "horodatage":    ts.isoformat(),
            "mesures":       mesures,
            "cycle_phase":   phase,
            "defaut_actif":  defaut_actif,
        }
        t_post = time.time()
        try:
            req = urllib.request.Request(
                url, data=json.dumps(body).encode("utf-8"),
                headers={"content-type": "application/json"}, method="POST",
            )
            with urllib.request.urlopen(req, timeout=2.0) as resp:
                last_status = f"HTTP {resp.status}"
                if resp.status in (200, 201):
                    n_ok += 1
                else:
                    n_err += 1
        except Exception as ex:
            n_err += 1
            last_status = f"{type(ex).__name__}: {ex}"
            if i < 5 or i % 60 == 0:
                print(f"  WARN post failed @ t={i}s : {ex}")
        latencies_ms.append((time.time() - t_post) * 1000.0)

        # Cadence temps reel
        target_t = (i + 1) / speed
        elapsed = time.time() - t_start
        if elapsed < target_t:
            time.sleep(target_t - elapsed)

    avg_lat = sum(latencies_ms) / max(1, len(latencies_ms))
    stats = {
        "n_total":       n,
        "n_ok":          n_ok,
        "n_err":         n_err,
        "avg_latency_ms": round(avg_lat, 2),
        "last_status":   last_status,
    }
    print(f"[post] termine: {n_ok}/{n} OK, latency_avg={avg_lat:.1f} ms, "
          f"last={last_status}")
    return stats


# =============================================================================
# 8. CLI
# =============================================================================
def main():
    p = argparse.ArgumentParser(description="VIMS Replay Simulator (CAT 994F1)")
    p.add_argument("--duration", type=float, default=600.0, help="Duree (s)")
    p.add_argument("--dt", type=float, default=1.0, help="Pas (s)")
    p.add_argument("--fault", choices=["", "ventilo_hs", "surchauffe_progressive",
                                       "fuite_huile", "niveau_bas"],
                   default="", help="Defaut a injecter")
    p.add_argument("--t-fault", type=float, default=60.0,
                   help="Instant d'apparition du defaut (s)")
    p.add_argument("--csv",  default="", help="Fichier CSV de sortie (1 Hz)")
    p.add_argument("--xlsx", default="", help="Fichier Excel format VIMS (snapshot 2 min)")
    p.add_argument("--post", default="", help="URL backend (envoi temps reel)")
    p.add_argument("--speed", type=float, default=1.0,
                   help="Multiplicateur vitesse (post seulement)")
    p.add_argument("--seed",  type=int, default=42, help="Seed du RNG")
    args = p.parse_args()

    if not (args.csv or args.xlsx or args.post):
        # Defaut : CSV nomme automatiquement
        args.csv = f"vims_replay_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    cfg = SimulatorConfig(
        duration_s=args.duration,
        dt=args.dt,
        fault=args.fault,
        t_fault=args.t_fault,
        speed=args.speed,
        seed=args.seed,
    )
    print(f"\n[VIMS replay] duree={cfg.duration_s}s  dt={cfg.dt}s  defaut={cfg.fault or 'aucun'}")

    timestamps, samples = simulate(cfg)
    sensors = load_sensors()

    if args.csv:
        export_csv(timestamps, samples, Path(args.csv), sensors)
    if args.xlsx:
        export_xlsx_vims_format(timestamps, samples, Path(args.xlsx), sensors)
    if args.post:
        post_realtime(timestamps, samples, args.post,
                      speed=args.speed, engin="994F1", fault=args.fault)

    print("\n[ok] Simulation terminee.")


if __name__ == "__main__":
    main()
